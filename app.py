import streamlit as st
import PyPDF2
import openpyxl
import pandas as pd
import docx
import os
import io
import time
import google.generativeai as genai
import subprocess
from io import StringIO
import sys
import matplotlib.pyplot as plt
from dotenv import load_dotenv
import concurrent.futures

# Load environment variables from .env file
load_dotenv()
GENAI_API_KEY = os.getenv('GOOGLE_API_KEY')

# Configure Google AI Platform API key
genai.configure(api_key=GENAI_API_KEY)

class CodeExecutionTimeout(Exception):
    """Custom exception for handling code execution timeouts."""
    pass

def read_pdf(file):
    """Read text content from a PDF file."""
    with io.BytesIO(file.read()) as f:
        reader = PyPDF2.PdfReader(f)
        text = ''.join(page.extract_text() for page in reader.pages)
    return text

def read_excel(file):
    """Read text content from an Excel file."""
    wb = openpyxl.load_workbook(file)
    text = ''
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                text += str(cell.value) + ' '
    return text

def read_csv(file):
    """Read text content from a CSV file."""
    df = pd.read_csv(file)
    return df.to_string(index=False)

def read_docx(file):
    """Read text content from a DOCX file."""
    doc = docx.Document(file)
    return '\n'.join(paragraph.text for paragraph in doc.paragraphs)

def generate_code(content, prompt, retries=3):
    """
    Generate Python code using the provided content and prompt.
    
    Parameters:
        content (str): The content to include in the prompt.
        prompt (str): The user-provided prompt for generating code.
        retries (int): Number of retries for generating code in case of failure.
    
    Returns:
        str: The generated Python code.
    """
    for attempt in range(retries):
        try:
            model = genai.GenerativeModel('gemini-1.0-pro-latest')
            combined_prompt = f"{content}\n{'Give python code to ' + prompt}"
            response = model.generate_content(combined_prompt)
            code = response.text.strip()
            if '...' in code:
                st.warning("Generated code is incomplete. Attempting to complete it...")
                code = complete_code(code, combined_prompt, content)
            return code
        except Exception as e:
            if attempt < retries - 1:
                st.warning(f"Error generating code: {e}. Retrying in 10 seconds... (Attempt {attempt + 1})")
                time.sleep(10)
            else:
                raise Exception(f"Failed to generate code after {retries} attempts.") from e

def complete_code(incomplete_code, previous_code, content):
    """
    Attempt to complete an incomplete piece of generated code.
    
    Parameters:
        incomplete_code (str): The incomplete code generated in the previous attempt.
        previous_code (str): The previously generated code.
        content (str): The content to include in the prompt.
    
    Returns:
        str: The completed Python code.
    """
    try:
        model = genai.GenerativeModel('gemini-1.0-pro-latest')
        combined_prompt = f"{previous_code}\n{incomplete_code}\n{content}"
        response = model.generate_content(combined_prompt)
        return response.text.strip()
    except Exception as e:
        raise Exception(f"Error completing code: {e}")

def validate_code(code):
    """
    Validate the generated Python code for potentially dangerous operations.
    
    Parameters:
        code (str): The Python code to validate.
    
    Returns:
        bool: True if the code is safe, False otherwise.
    """
    unsafe_keywords = ['import os', 'import sys', 'subprocess', 'eval', 'exec']
    for keyword in unsafe_keywords:
        if keyword in code:
            return False
    return True

def execute_code_with_timeout(code, globals_dict, timeout=10):
    """
    Execute the code with a timeout.

    Parameters:
        code (str): The code to execute.
        globals_dict (dict): The globals dictionary for exec().
        timeout (int): The timeout in seconds.

    Returns:
        str: The captured standard output from the code execution.
    """
    def exec_code():
        exec(code, globals_dict)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
    future = executor.submit(exec_code)

    try:
        future.result(timeout=timeout)
    except concurrent.futures.TimeoutError:
        raise CodeExecutionTimeout("Code execution timed out.")

def execute_code(code, content):
    """
    Execute the generated Python code and handle the output.
    
    Parameters:
        code (str): The Python code to execute.
        content (str): The content to be used for replacing file inputs in the code.
    """
    if not validate_code(code):
        st.error("Generated code contains unsafe operations and cannot be executed.")
        return

    try:
        code = code.replace('pd.read_csv', f"pd.read_csv(io.StringIO('''{content}'''))")
        code = code.replace('pd.read_excel', f"pd.read_excel(io.BytesIO('''{content}'''))")
        code = code.replace('open', f"io.BytesIO('''{content}''')")

        # Redirect standard output to a StringIO buffer
        stdout = sys.stdout
        sys.stdout = StringIO()

        # Extracting the code block between ```python tags
        if "```python" in code:
            code = code.split("```python")[1].strip()
            code = code.split("```")[0].strip()

        exec_globals = {'io': io, 'pd': pd, 'plt': plt}

        # Execute the code with a timeout
        execute_code_with_timeout(code, exec_globals)

        # Retrieve the captured standard output
        output = sys.stdout.getvalue()
        sys.stdout.close()
        sys.stdout = stdout

        # Check if the output is a diagram
        if 'plt.show()' in code:
            st.pyplot(plt)
        else:
            st.write("Output:")
            st.write(output.strip())

        # Show balloons animation upon successful execution
        st.balloons()

    except CodeExecutionTimeout:
        st.error("Code execution timed out. Please check your code for infinite loops or long-running operations.")
    except ModuleNotFoundError as e:
        missing_module = str(e).split("'")[1]
        try:
            subprocess.run(['pip', 'install', '--user', missing_module], check=True)
            execute_code_with_timeout(code, exec_globals)
            output = sys.stdout.getvalue()
            sys.stdout.close()
            sys.stdout = stdout

            if 'plt.show()' in code:
                st.pyplot(plt)
            else:
                st.write("Output:")
                st.write(output.strip())
        except Exception as install_error:
            st.error(f"Error installing missing module '{missing_module}': {install_error}")
    except Exception as e:
        st.error(f"Error executing code: {e}")

def main():
    """
    Main function to run the Streamlit application.
    """
    st.set_page_config(page_title="AI Code Generator & Interpreter", page_icon=":robot_face:", layout="wide")
    st.title("Code Interpreter")
    st.subheader("Upload a file and provide a prompt to generate and execute Python code.")

    with st.sidebar:
        st.header("Upload and Prompt")
        file = st.file_uploader("Upload file")
        prompt = st.sidebar.text_input("Input prompt", placeholder="Provide a prompt")

    if file:
        _, file_extension = os.path.splitext(file.name)
        file_extension = file_extension.lower()

        if file_extension == '.pdf':
            content = read_pdf(file)
        elif file_extension == '.xlsx':
            content = read_excel(file)
        elif file_extension == '.csv':
            content = read_csv(file)
        elif file_extension == '.docx':
            content = read_docx(file)
        else:
            st.error("Unsupported file format")
            return

        if content:
            if st.sidebar.button("Generate Code"):
                try:
                    with st.spinner("Generating code..."):
                        code = generate_code(content, prompt)
                        st.code(code, language='python')
                    with st.spinner("Executing code..."):
                        execute_code(code, content)
                except SyntaxError as e:
                    st.error(f"Syntax Error: {e}")
                    try:
                        corrected_code = generate_code(content, prompt, retries=1)
                        st.warning("Attempting to correct the code...")
                        st.code(corrected_code, language='python')
                        st.write("Executing corrected code...")
                        execute_code(corrected_code, content)
                    except Exception as e:
                        st.error(f"Failed to correct code: {e}")
                except Exception as e:
                    st.error(f"Error: {e}")

if __name__ == '__main__':
    main()
