import streamlit as st
import torch
from transformers import AutoTokenizer, CodeGenModel
import PyPDF2
import openpyxl
import pandas as pd
import docx
import io
import os
import subprocess

# Load the pre-trained model and tokenizer
model_name = "Salesforce/codegen-2B-mono"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = CodeGenModel.from_pretrained(model_name)

# Define the functions to read different file types
def read_pdf(file):
    text = ''
    with io.BytesIO(file.read()) as f:
        reader = PyPDF2.PdfReader(f)
        for page_num in range(len(reader.pages)):
            text += reader.pages[page_num].extract_text()
    return text

def read_excel(file):
    wb = openpyxl.load_workbook(file)
    sheets = wb.sheetnames
    text = ''
    for sheet_name in sheets:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                text += str(cell.value) + ' '
    return text

def read_csv(file):
    df = pd.read_csv(file)
    text = df.to_string(index=False)
    return text

def read_docx(file):
    doc = docx.Document(file)
    text = ''
    for paragraph in doc.paragraphs:
        text += paragraph.text + '\n'
    return text

# Define the function to generate code using the Hugging Face API
def generate_code(content, prompt):
    inputs = tokenizer(prompt, return_tensors="pt")
    inputs["input_ids"] = inputs["input_ids"].flatten()
    inputs["attention_mask"] = inputs["attention_mask"].flatten()

    chunk_size = 1024  # Adjust this value based on your memory requirements
    chunks = [content[i:i + chunk_size] for i in range(0, len(content), chunk_size)]

    generated_code = []
    for chunk in chunks:
        outputs = model.generate(**inputs, max_length=256, num_beams=2, early_stopping=True, batch_size=1)
        code = tokenizer.decode(outputs[0], skip_special_tokens=True)
        generated_code.append(code)

    return '\n'.join(generated_code)

# Define the function to execute the generated code
def execute_code(code):
    # Create a temporary file to store the generated code
    temp_file = 'temp.py'
    with open(temp_file, 'w') as f:
        f.write(code)

    # Execute the generated code using subprocess
    try:
        result = subprocess.check_output(['python', temp_file])
        return result.decode('utf-8')
    except subprocess.CalledProcessError as e:
        return str(e)
    finally:
        # Remove the temporary file
        os.remove(temp_file)

# Main function
def main():
    st.title("File Interpreter")

    file = st.file_uploader("Upload file")

    if file is not None:
        # Get file extension
        _, file_extension = os.path.splitext(file.name)
        file_extension = file_extension.lower()

        if file_extension == '.pdf':
            content = read_pdf(file)
        elif file_extension == '.xlsx':
            content = read_excel(file)
        elif file_extension == '.csv':
            content = read_csv(file)
        elif file_extension == '.docx':
            content = read_docx(file)
        else:
            st.error("Unsupported file format")
            return

        if content:
            prompt = st.text_input("Input prompt", "Provide a prompt")
            if st.button("Generate Code"):
                code = generate_code(content, prompt)
                result = execute_code(code)
                st.write(result)

if __name__ == "__main__":
    main()